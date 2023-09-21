from openpyxl import load_workbook, Workbook


lista_arquivos = ["CustosAutom", "PopulacaoPOA", "SuperMercado"]


wb = Workbook()
nome_arquivo_final = "resultados.xlsx"

for nome_arquivo in lista_arquivos:
     arquivo = load_workbook(filename="%s.xlsx" % nome_arquivo)
     sheet = arquivo[nome_arquivo]
     max_linhas = sheet.max_row
     max_colunas = sheet.max_column

     ws = wb.create_sheet(title=nome_arquivo)

     for i in range(1, max_linhas +1):
         for j in range(1, max_colunas + 1):
             celula = sheet.cell(row=i, column=j)

             ws.cell(row=i, column=j).value = celula.value

wb.remove(wb["Sheet"])
wb.save(filename=nome_arquivo_final)
