from openpyxl import load_workbook

wb =load_workbook(filename="nomes.xlsx")

planilha = wb['Planilha1']

for i in range(2, 5):
    nome = planilha[f'A{i}'].value
    idade = planilha[f'B{i}'].value
    print(f"{nome} tem {idade} anos.")

planilha['B5'] = "=SUM(B2:B4)"

planilha['A7'] = 'ALUNOS'
planilha.merge_cells("A7:B7")
planilha.unmerge_cells("A7:B7")

wb.save(filename="nomes2.xlsx")
